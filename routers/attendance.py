"""
Attendance router - list records, stats, delete, export reports, and email notifications.
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, distinct
from typing import List, Optional
from datetime import date, timedelta, datetime
from io import BytesIO
import smtplib
import os
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from dotenv import load_dotenv

import models
import schemas
from database import get_db

load_dotenv()

router = APIRouter(prefix="/attendance", tags=["Attendance"])

# ─── Email Config ────────────────────────────────────────────────────────────
EMAIL_HOST = os.getenv("EMAIL_HOST", "smtp.gmail.com")
EMAIL_USER = os.getenv("EMAIL_HOST_USER", "pawantripathi802@gmail.com")
EMAIL_PASS = os.getenv("EMAIL_HOST_PASSWORD", "kplopuunmezfesie")
EMAIL_FROM = f"Face Recognition System <{EMAIL_USER}>"


def _send_attendance_email(student: models.Student, attendance: models.Attendance):
    """
    Send an attendance confirmation email via Gmail SMTP SSL (port 465).
    Returns (True, '') on success or (False, error_message) on failure.
    Uses SSL directly — more reliable than STARTTLS from cloud providers like Render.
    """
    import ssl
    if not student.email:
        return False, "Student has no email address"
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Attendance Marked - {attendance.date}"
        msg["From"] = EMAIL_FROM
        msg["To"] = student.email

        html_body = f"""
        <html><body style="font-family: Arial, sans-serif; background:#f4f4f4; padding:20px;">
        <div style="max-width:600px;margin:auto;background:#fff;border-radius:8px;padding:30px;box-shadow:0 2px 8px rgba(0,0,0,0.1);">
            <h2 style="color:#667eea;">Attendance Marked</h2>
            <p>Dear <strong>{student.name}</strong>,</p>
            <p>Your attendance has been successfully recorded.</p>
            <table style="width:100%;border-collapse:collapse;margin:20px 0;">
                <tr style="background:#f7fafc;"><td style="padding:10px;font-weight:bold;">Student ID</td><td style="padding:10px;">{student.student_id}</td></tr>
                <tr><td style="padding:10px;font-weight:bold;">Department</td><td style="padding:10px;">{student.get_department_display()}</td></tr>
                <tr style="background:#f7fafc;"><td style="padding:10px;font-weight:bold;">Date</td><td style="padding:10px;">{attendance.date}</td></tr>
                <tr><td style="padding:10px;font-weight:bold;">Time</td><td style="padding:10px;">{attendance.time}</td></tr>
                <tr style="background:#f7fafc;"><td style="padding:10px;font-weight:bold;">Status</td><td style="padding:10px;color:green;"><strong>{attendance.get_status_display()}</strong></td></tr>
                <tr><td style="padding:10px;font-weight:bold;">Confidence</td><td style="padding:10px;">{attendance.confidence:.1f}%</td></tr>
            </table>
            <p style="color:#718096;font-size:13px;">This is an automated message from the Face Recognition Attendance System.</p>
        </div></body></html>
        """
        msg.attach(MIMEText(html_body, "html"))

        # Use SMTP_SSL port 465 — works reliably from cloud servers (Render, Heroku, etc.)
        ssl_ctx = ssl.create_default_context()
        with smtplib.SMTP_SSL(EMAIL_HOST, 465, context=ssl_ctx, timeout=20) as server:
            server.login(EMAIL_USER, EMAIL_PASS)
            server.sendmail(EMAIL_USER, student.email, msg.as_string())
        return True, ""

    except smtplib.SMTPAuthenticationError as e:
        err = f"Gmail authentication failed: {e.smtp_error.decode() if hasattr(e, 'smtp_error') else str(e)}"
        print(f"Email auth error: {err}")
        return False, err
    except Exception as e:
        err = f"{type(e).__name__}: {e}"
        print(f"Email error for {student.email}: {err}")
        return False, err


# ─── Helpers ─────────────────────────────────────────────────────────────────

def build_filtered_query(db: Session, start_date=None, end_date=None,
                          department=None, course=None, semester=None):
    """Build a filtered attendance query."""
    query = db.query(models.Attendance).join(models.Student)
    if start_date:
        query = query.filter(models.Attendance.date >= start_date)
    if end_date:
        query = query.filter(models.Attendance.date <= end_date)
    if department:
        query = query.filter(models.Student.department == department)
    if course:
        query = query.filter(models.Student.course == course)
    if semester:
        query = query.filter(models.Student.semester == semester)
    return query


def att_to_out(att: models.Attendance) -> schemas.AttendanceOut:
    return schemas.AttendanceOut(
        id=att.id,
        student_id=att.student_id,
        student_name=att.student.name if att.student else None,
        student_uid=att.student.student_id if att.student else None,
        date=att.date,
        time=att.time,
        status=att.status,
        status_display=att.get_status_display(),
        confidence=att.confidence,
    )


# ─── Endpoints ───────────────────────────────────────────────────────────────

@router.get("/", response_model=List[schemas.AttendanceOut], summary="List attendance records")
def list_attendance(
    start_date: Optional[str] = Query(None, description="Filter from date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="Filter to date (YYYY-MM-DD)"),
    department: Optional[str] = Query(None),
    course: Optional[str] = Query(None),
    semester: Optional[str] = Query(None),
    limit: int = Query(200, le=1000),
    db: Session = Depends(get_db)
):
    """Get all attendance records, newest first. Supports optional date and filter params."""
    query = build_filtered_query(db, start_date, end_date, department, course, semester)
    records = query.order_by(models.Attendance.date.desc(), models.Attendance.time.desc()).limit(limit).all()
    return [att_to_out(a) for a in records]


@router.get("/stats", response_model=schemas.AttendanceStatsOut, summary="Dashboard statistics")
def get_stats(db: Session = Depends(get_db)):
    """Returns overall statistics for the attendance dashboard."""
    today_str = date.today().isoformat()
    week_start = (date.today() - timedelta(days=6)).isoformat()

    total_students = db.query(func.count(models.Student.id)).scalar() or 0
    total_records = db.query(func.count(models.Attendance.id)).scalar() or 0
    today_attendance = db.query(func.count(models.Attendance.id)).filter(
        models.Attendance.date == today_str
    ).scalar() or 0
    today_percentage = round((today_attendance / total_students * 100) if total_students > 0 else 0.0, 1)
    week_attendance = db.query(func.count(models.Attendance.id)).filter(
        models.Attendance.date >= week_start,
        models.Attendance.date <= today_str
    ).scalar() or 0
    avg_confidence = db.query(func.avg(models.Attendance.confidence)).scalar() or 0.0

    # Department breakdown (for today)
    dept_rows = db.query(
        models.Student.department,
        func.count(distinct(models.Student.id)).label('total'),
        func.count(models.Attendance.id).label('present_today')
    ).outerjoin(
        models.Attendance,
        (models.Attendance.student_id == models.Student.id) & (models.Attendance.date == today_str)
    ).group_by(models.Student.department).all()

    dept_stats = []
    dept_map = models.Student.DEPARTMENT_CHOICES
    for row in dept_rows:
        dept_stats.append({
            'department': row.department,
            'department_display': dept_map.get(row.department, row.department),
            'total': row.total,
            'present_today': row.present_today
        })

    return schemas.AttendanceStatsOut(
        total_students=total_students,
        total_records=total_records,
        today_attendance=today_attendance,
        today_percentage=today_percentage,
        week_attendance=week_attendance,
        avg_confidence=round(avg_confidence, 1),
        dept_stats=dept_stats
    )


@router.delete("/{attendance_id}", summary="Delete an attendance record")
def delete_attendance(attendance_id: int, db: Session = Depends(get_db)):
    """Delete a specific attendance record by its ID."""
    record = db.query(models.Attendance).filter(models.Attendance.id == attendance_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Attendance record not found.")
    db.delete(record)
    db.commit()
    return {"success": True, "message": f"Attendance record {attendance_id} deleted."}


# ─── Export Endpoints ─────────────────────────────────────────────────────────

@router.get("/export/excel", summary="Download Excel attendance report")
def export_excel(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    department: Optional[str] = Query(None),
    course: Optional[str] = Query(None),
    semester: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Generate and download an Excel (.xlsx) attendance report."""
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment

        query = build_filtered_query(db, start_date, end_date, department, course, semester)
        records = query.order_by(models.Attendance.date.desc()).all()

        data = []
        for att in records:
            data.append({
                'Student ID': att.student.student_id,
                'Name': att.student.name,
                'Department': att.student.get_department_display(),
                'Course': att.student.get_course_display(),
                'Semester': f'Sem {att.student.semester}',
                'Date': att.date,
                'Time': att.time,
                'Status': att.get_status_display(),
                'Confidence': f'{att.confidence:.1f}%'
            })

        df = pd.DataFrame(data)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Attendance', index=False)
            ws = writer.sheets['Attendance']
            for cell in ws[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        output.seek(0)
        filename = f'attendance_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel: {str(e)}")


@router.get("/export/pdf", summary="Download PDF attendance report")
def export_pdf(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    department: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Generate and download a PDF attendance report (max 100 records)."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER

        query = build_filtered_query(db, start_date, end_date, department)
        records = query.order_by(models.Attendance.date.desc()).limit(100).all()

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title', parent=styles['Heading1'], fontSize=24,
            textColor=colors.HexColor('#667eea'), alignment=TA_CENTER, spaceAfter=20
        )
        elements = [
            Paragraph("Attendance Report", title_style),
            Spacer(1, 0.2 * inch),
            Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']),
            Spacer(1, 0.3 * inch),
        ]

        table_data = [['Student ID', 'Name', 'Department', 'Date', 'Time', 'Confidence']]
        for att in records:
            table_data.append([
                att.student.student_id,
                att.student.name,
                att.student.get_department_display(),
                att.date,
                att.time[:5],
                f'{att.confidence:.1f}%'
            ])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')]),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        filename = f'attendance_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        return StreamingResponse(
            buffer,
            media_type='application/pdf',
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")


@router.get("/export/student-pdf/{student_id}", summary="Download per-student attendance PDF")
def export_student_pdf(student_id: str, db: Session = Depends(get_db)):
    """Generate and download a PDF with all attendance records for a specific student."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER

        student = db.query(models.Student).filter(models.Student.student_id == student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail="Student not found.")

        records = db.query(models.Attendance).filter(
            models.Attendance.student_id == student.id
        ).order_by(models.Attendance.date.desc()).all()

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title', parent=styles['Heading1'], fontSize=22,
            textColor=colors.HexColor('#667eea'), alignment=TA_CENTER, spaceAfter=20
        )
        elements = [
            Paragraph("Student Attendance Record", title_style),
            Spacer(1, 0.2 * inch),
        ]

        # Student info table
        info_data = [
            ['Student ID:', student.student_id],
            ['Name:', student.name],
            ['Department:', student.get_department_display()],
            ['Course:', student.get_course_display()],
            ['Semester:', f'Semester {student.semester}'],
            ['Division:', student.division],
        ]
        info_table = Table(info_data, colWidths=[2 * inch, 4 * inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph(f"<b>Total Records:</b> {len(records)} | <b>Generated:</b> {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
        elements.append(Spacer(1, 0.2 * inch))

        if records:
            att_data = [['Date', 'Time', 'Status', 'Confidence']]
            for att in records:
                att_data.append([att.date, att.time[:5], att.get_status_display(), f'{att.confidence:.1f}%'])
            att_table = Table(att_data, repeatRows=1)
            att_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')]),
            ]))
            elements.append(att_table)

        doc.build(elements)
        buffer.seek(0)
        filename = f'attendance_{student.student_id}_{date.today().isoformat()}.pdf'
        return StreamingResponse(
            buffer,
            media_type='application/pdf',
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")


# ─── Email Endpoints ──────────────────────────────────────────────────────────

@router.post("/{attendance_id}/send-email", summary="Send attendance email for a specific record")
def send_email_for_record(attendance_id: int, db: Session = Depends(get_db)):
    """Send an HTML attendance confirmation email to the student for a specific attendance record."""
    record = db.query(models.Attendance).filter(models.Attendance.id == attendance_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Attendance record not found.")

    student = record.student
    if not student.email:
        return {"success": False, "message": f"Student {student.student_id} has no email address."}

    ok, err_msg = _send_attendance_email(student, record)
    if ok:
        return {"success": True, "message": f"Email sent successfully to {student.email}."}
    else:
        return {"success": False, "message": f"Failed to send email to {student.email}. Error: {err_msg}"}


@router.post("/send-all-emails", summary="Send attendance emails to all records")
def send_all_emails(db: Session = Depends(get_db)):
    """
    Send attendance emails to every attendance record in the database.
    Returns counts of successful, failed, and skipped (no email) records.
    """
    records = db.query(models.Attendance).join(models.Student).all()

    success_count = 0
    failed_count = 0
    no_email_count = 0

    for record in records:
        student = record.student
        if not student.email:
            no_email_count += 1
            continue
        try:
            ok, _ = _send_attendance_email(student, record)
            if ok:
                success_count += 1
            else:
                failed_count += 1
        except Exception:
            failed_count += 1

    return {
        "success": True,
        "sent": success_count,
        "failed": failed_count,
        "skipped_no_email": no_email_count,
        "message": f"Sent {success_count} email(s). Failed: {failed_count}. Skipped (no email): {no_email_count}."
    }

